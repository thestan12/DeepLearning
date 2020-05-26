
import matplotlib.pyplot as plt
import openpyxl as op
import datetime
import uuid

imgPath = "./image/"
excelColName = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

def getRamdomFileName():
    return str(uuid.uuid4()) + ".png";

def display_img(img):
    filename = getRamdomFileName()
    plt.figure()
    plt.imshow(img)
    plt.colorbar()
    plt.grid(False)
    plt.subplots_adjust(left=0, right=1, top=0.9, bottom=0.1)
    plt.savefig(imgPath + filename, dpi = 150)
    return filename;


def addRow(tab, imgTab):
    wb = op.load_workbook("./results.xlsx", read_only=False)
    ws = wb.active

    id = ws.max_row
    if id != 1 or ws["A1"].value is not None:
        id += 1

    tab.insert(0, id)
    ws.append(tab)

    i = 1
    for f in imgTab:
        img = op.drawing.image.Image(imgPath + f)
        # img.anchor = excelColName[len(tab)] + str(id)
        img.anchor = ws.cell(row=id, column=len(tab)+i).coordinate
        ws.add_image(img)
        i += 1

    wb.save('./results.xlsx')

def display_accuracy_Fig(logs):
    plt.plot(logs.history['accuracy'])
    plt.plot(logs.history['val_accuracy'])
    plt.show()

def display_accuracy_Fig(logs):
    plt.plot(logs.history['accuracy'])
    plt.plot(logs.history['val_accuracy'])
    plt.show()
